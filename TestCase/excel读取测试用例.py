# 1、制定解析规则
# 从excel加载复杂的数据结构，可以变为：dict, json, yaml
# -*- coding:utf-8 -*-
# 读取Excel数据
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
def read_excel(xls_path):
    wb = load_workbook(xls_path)
    ws = wb.active  # 获取当前的工作表
    return ws

re = read_excel('../data/test_excel.xlsx')
# Excel内容格式
# 第一行：表头
# 第x行：（用例ID为0）测试套件的信息
# 第x行：（步骤ID为0）测试用例的信息
# 第x行：测试执行步骤（关键字）
'''
    重构目标：
        1. 封装为函数，函数返回处理后的数据。
        2. 清理垃圾数据，移除没用的None
        3. 让数据更有可读性。
'''

def handle_excel(ws: Worksheet):
    _suite = {  # 套件，本工作表中所有的测试用例
        'info': {},  # 套件的属性信息
        'case_list': []  # 用例列表
    }
    _case_list = _suite['case_list']
    _case = {  # 用例
        'info': {},  # 用例的属性信息
        'steps': []  # 本用例的所有步骤
    }
    top_keys = [i for i in list(re.iter_rows(max_row=1, values_only=True))[0] if i is not None]
    print(top_keys)
    # 使用高阶函数filter过滤垃圾数据
    top_keys = filter(
        lambda x: x is not None,
        list(ws.iter_rows(max_row=1, values_only=True))[0])
    print(list(top_keys))

    # 使用python解包
    for (
            用例,
            步骤,
            步骤名,
            关键字,
            参数1,
            参数2,
            执行结果,
            报告摘要
    ) in ws.iter_rows(min_row=2, values_only=True):
        if 用例 == 0:  # 用例id==0，是分组信息
            _suite['info'][步骤名] = 参数1
        elif 步骤 == 0:  # 步骤id==0，是用例信息
            _case = {
                'info': {},
                'steps': [],
            }
            _suite['case_list'].append(_case)
            _case['info'][步骤名] = 参数1
        else:
            _case['steps'].append(
                {
                    '步骤名': 步骤名,
                    '关键字': 关键字,
                    '参数1': list(
                        filter(
                            lambda x: x is not None,
                            [
                                参数1,
                                参数2
                            ]
                        )
                    )
                }
            )
    return _suite
print(handle_excel(re))
'''
data = list(re.values)
data = data[1:]     # 干掉第一行：切片
for d in data:  # 逐行打印 >> 打包数据
    if d[0] == 0:
        # 这是分组信息

        print('suite info:', d)

    elif d[1] == 0 :

        # 这是用例信息

        print('  test case info:', d)

    else:

    	# 这是用例步骤

        print('    step：', d)



new_data = {}   # 数据打包之后，传递给测试框架，创建测试用例并执行

'''
