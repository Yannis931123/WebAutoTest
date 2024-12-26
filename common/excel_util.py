from pathlib import Path
import openpyxl


class ExcelUtil(object):
    def __init__(self):
        # 项目目录
        self.base_dir = Path(__file__).parent.parent

    # openpyxl（更多用这个）、xlrd
    def read_excel(self):
        # 第一步：加载文件
        wb = openpyxl.load_workbook('%s/data/test_excel.xlsx' % self.base_dir)
        # 第二步：获得sheet对象
        sheet = wb['login']
        # 第三步：获得excel的行数和列数
        # print(sheet.max_row, sheet.max_column)
        all_list = []
        for rows in range(2, sheet.max_row+1):  # 第二行开始，去掉行标题
            temp_list = []
            for cols in range(1, sheet.max_column+1):
                # print(sheet.cell(rows, cols).value)
                temp_list.append(sheet.cell(rows, cols).value)
            # print(temp_list)
            all_list.append(temp_list)

        # print(all_list)
        return all_list



if __name__ == '__main__':
    ExcelUtil().read_excel()
    print(ExcelUtil().read_excel())

